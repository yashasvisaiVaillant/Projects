import os
import argparse
import pandas as pd
from bs4 import BeautifulSoup
import chardet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    parser = argparse.ArgumentParser(description='Generate failure report from HTML report file.')
    parser.add_argument('--html_file', type=str, required=True, help='Path to the input HTML report file.')
    parser.add_argument('--output_file', type=str, default='failure_report.xlsx', help='Path for the output Excel file.')
    args = parser.parse_args()

    html_file = args.html_file
    output_file = args.output_file

    # Extract the directory name containing the html_file
    test_name = os.path.basename(os.path.dirname(html_file))
    print(test_name)

    # Detect file encoding
    with open(html_file, 'rb') as file:
        rawdata = file.read()
    result = chardet.detect(rawdata)
    encoding = result['encoding']

    # Read HTML content
    with open(html_file, encoding=encoding) as file:
        html_content = file.read()

    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all 'Failed' divs
    failed_decisions = soup.find_all('div', class_='Failed')

    # Prepare data
    data = []
    concatenated_list = []

    # Skip first 4, process rest
    for div in failed_decisions[4:]:
        header_div = div.find('div', class_='header_Failed')
        if header_div:
            # Extract tb_name
            tb_name_span = header_div.find('span', class_='tb_name')
            tb_name = tb_name_span.get_text(strip=True) if tb_name_span else None

            # Extract tb_library
            tb_library_span = header_div.find('span', class_='tb_library')
            tb_library = tb_library_span.get_text(strip=True) if tb_library_span else None

            # Append concatenated string
            if tb_name and tb_library:
                concatenated_list.append(f"{tb_name}.{tb_library}")
            elif tb_name:
                concatenated_list.append(f"{tb_name}")
            elif tb_library:
                concatenated_list.append(f"{tb_library}")
        else:
            continue

    reason = ' :: '.join(concatenated_list)
    data.append({'Test Name': test_name, 'Failure Reason': reason})

    # Create DataFrame
    df = pd.DataFrame(data)

    # Save to Excel
    df.to_excel(output_file, index=False)

    # Format the Excel file
    wb = load_workbook(output_file)
    ws = wb.active

    # Adjust column widths and wrap text
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            # Enable wrap_text if desired
            # cell.alignment = Alignment(wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"Formatted Excel report saved to {output_file}")

if __name__ == '__main__':
    main()