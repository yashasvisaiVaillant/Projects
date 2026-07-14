from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class ExcelFormatter:
    def __init__(self, filename):
        self.filename = filename

    def adjust_columns(self):
        wb = load_workbook(self.filename)
        ws = wb.active

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                # Enable wrap_text for better readability
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Add some padding to width
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(self.filename)